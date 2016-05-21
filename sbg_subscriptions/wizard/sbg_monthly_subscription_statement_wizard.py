# -*- coding: utf-8 -*-
from openerp import models, fields, api, _
from datetime import date, datetime
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.worksheet import ColumnDimension
from openpyxl.styles import Border, Side, Font, Alignment
import base64
from dateutil.parser import parse
from dateutil.relativedelta import relativedelta
from calendar import monthrange

class sbg_monthly_subscription_statement_wizard(models.TransientModel):
    _name = 'sbg.monthly.subscription.statement.wizard'
    _description = 'Monthly subscription statement wizard'

    def _get_default_service(self):
        return self.env['sbg.subscription.services'].search([])[0]

    def _start_of_year(self):
        now = date(date.today().year, 1, 1)
        return now

    def _end_of_year(self):
        now = date(date.today().year, 12, 31)

    def get_date(self, tup):
        return tup.date

    subscription_service_id = fields.Many2one('sbg.subscription.services', 'Subscription service', default=_get_default_service)
    start_date = fields.Date(string="Start date", default=_start_of_year)
    end_date = fields.Date(string="End date", default=_end_of_year)
    name = fields.Char('File Name', readonly=True)
    data = fields.Binary('File', readonly=True)
    state = fields.Selection([('choose', 'choose'),  # choose subscription and period
                              ('get', 'get')],       # get the file
                              default='choose')

    @api.onchange('subscription_service_id')
    def onchange_total_amount(self):
        self.start_date = self.subscription_service_id.start_date
        if self.subscription_service_id.duration_type == 'permanent':
            self.end_date = date(date.today().year, 12, 31)
        else:
            self.end_date = self.subscription_service_id.end_date

    @api.multi
    def generate_file(self, context=None):

        def _get_column_letter(col_idx):
            """Convert a column number into a column letter (3 -> 'C')

            Right shift the column col_idx by 26 to find column letters in reverse
            order.  These numbers are 1-based, and can be converted to ASCII
            ordinals by adding 64.

            """
            # these indicies corrospond to A -> ZZZ and include all allowed
            # columns
            if not 1 <= col_idx <= 18278:
                raise ValueError("Invalid column index {0}".format(col_idx))
            letters = []
            while col_idx > 0:
                col_idx, remainder = divmod(col_idx, 26)
                # check for exact division and borrow if needed
                if remainder == 0:
                    remainder = 26
                    col_idx -= 1
                letters.append(chr(remainder + 64))
            return ''.join(reversed(letters))

        _COL_STRING_CACHE = {}
        _STRING_COL_CACHE = {}
        for col_cache in range(1, 18279):
            col_value = _get_column_letter(col_cache)
            _STRING_COL_CACHE[col_cache] = col_value
            _COL_STRING_CACHE[col_cache] = col_value

        def get_column_letter(idx, ):
            """Convert a column index into a column letter
            (3 -> 'C')
            """
            try:
                return _STRING_COL_CACHE[idx]
            except KeyError:
                raise ValueError("Invalid column index {0}".format(idx))

        start_date = parse(self.start_date)
        end_date = parse(self.end_date)

        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()

        border_top = Border(top=Side(style='thin'))
        border_bottom = Border(bottom=Side(style='thin'))
        border_left = Border(left=Side(style='thin'))
        border_right = Border(right=Side(style='thin'))
        border_top_left = Border(top=Side(style='thin'), left=Side(style='thin'))
        border_top_right = Border(top=Side(style='thin'), right=Side(style='thin'))
        border_bottom_left = Border(bottom=Side(style='thin'), left=Side(style='thin'))
        border_bottom_right = Border(bottom=Side(style='thin'), right=Side(style='thin'))
        font_bold = Font(bold=True)
        font_h1 = Font(size=18, bold=True)
        font_h2 = Font(size=14, bold=True)

        wb = Workbook()

        ws = wb.active
        ws.title = _('Monthly statement')
        row = 1
        ws.cell(row=row, column=1).value = self.subscription_service_id.name
        ws.cell(row=row, column=1).font = font_h1
        ws.column_dimensions['A'].width = 15

        row += 1
        ws.cell(row=row, column=1).value = _('Period')
        ws.cell(row=row, column=1).font = font_h2
        ws.cell(row=row, column=2).value = start_date
        ws.cell(row=row, column=2).number_format = 'dd-mmm-yyyy'
        ws.cell(row=row, column=3).value = end_date
        ws.cell(row=row, column=3).number_format = 'dd-mmm-yyyy'

        row += 2
        ws.cell(row=row, column=1).border = border_right

        row += 1
        ws.cell(row=row, column=1).value = _('Subscriber')
        ws.cell(row=row, column=1).font = font_bold
        ws.cell(row=row, column=1).border = border_bottom
        ws.cell(row=row, column=1).border = border_bottom_right

        start_dates = []
        end_dates = []
        col = 2

        date = start_date
        while date < end_date:
            month_end = datetime(date.year, date.month, monthrange(date.year, date.month)[1])
            start_dates.append(date)
            end_dates.append(month_end)
            ws.cell(row=row-1, column=col).value = month_end
            ws.cell(row=row-1, column=col).font = font_bold
            ws.cell(row=row-1, column=col).number_format = 'mmmm-yyyy'
            ws.cell(row=row-1, column=col).alignment = Alignment(horizontal='center')
            ws.merge_cells(get_column_letter(col) + str(row-1) + ':' + get_column_letter(col+1) + str(row-1))
            ws.cell(row=row, column=col).value = _('Debits')
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=col).border = border_bottom
            ws.column_dimensions[get_column_letter(col)].width = 12
            ws.cell(row=row, column=col+1).value = _('Credits')
            ws.cell(row=row, column=col+1).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=col+1).border = border_bottom
            ws.column_dimensions[get_column_letter(col+1)].width = 12
            col += 2
            date = month_end + relativedelta(days=1)
        ws.cell(row=row-1, column=col).value = _('Total')
        ws.cell(row=row-1, column=col).alignment = Alignment(horizontal='right')
        ws.cell(row=row-1, column=col).font = font_bold
        ws.cell(row=row-1, column=col).border = border_left
        ws.cell(row=row-1, column=col + 1).value = _('Total')
        ws.cell(row=row-1, column=col + 1).alignment = Alignment(horizontal='right')
        ws.cell(row=row-1, column=col + 1).font = font_bold
        ws.cell(row=row, column=col).value = _('Debits')
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=col).font = font_bold
        ws.cell(row=row, column=col).border = border_bottom_left
        ws.cell(row=row, column=col + 1).value = _('Credits')
        ws.cell(row=row, column=col + 1).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=col + 1).font = font_bold
        ws.cell(row=row, column=col + 1).border = border_bottom
        ws.cell(row=row, column=col + 2).value = _('Difference')
        ws.cell(row=row, column=col + 2).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=col + 2).font = font_bold
        ws.cell(row=row, column=col + 2).border = border_bottom
        total_col = col
        first_row = row + 1

        #
        # Search selected subscription and included products
        #
        product_ids = []
        for product in self.subscription_service_id.product_ids:
            if product.id not in product_ids:
                product_ids.append(product.id)
        #
        # Get debits from subscription statement model query
        #
        sql_debits = """
            SELECT r.id, r.display_name, t.date, t.value
            FROM sbg_subscriptions s
            JOIN res_partner r
            ON r.id = s.partner_id
            JOIN sbg_subscription_statement t
            ON t.subscription_id = s.id
            AND t.date BETWEEN %s AND %s
            WHERE s.subscription_service_id = %s
            AND s.start_date <= %s
            ORDER BY r.display_name
        """

        #
        # Get credits from subscription statement model query
        #
        sql_credits = """
            SELECT sum(l.price_unit * l.quantity) As value
            FROM account_invoice i
            JOIN account_invoice_line l
            ON l.invoice_id = i.id
            AND l.product_id in %s
            WHERE i.partner_id = %s
            AND i.date_invoice between %s and %s
            AND i.state = 'paid'
        """
        self.env.cr.execute(sql_debits, (self.start_date, self.end_date, self.subscription_service_id.id, self.end_date,))
        last_subscriber = ''
        subscribers = self.env.cr.dictfetchall()
        for subscriber in subscribers:
            if subscriber['display_name'] != last_subscriber:
                last_subscriber = subscriber['display_name']
                row += 1
                ws.cell(row=row, column=1).value = subscriber['display_name']
                ws.cell(row=row, column=2).border = border_left
                ws.cell(row=row, column=total_col).value = '='
                ws.cell(row=row, column=total_col).font = font_bold
                ws.cell(row=row, column=total_col).number_format = '#,##0.00'
                ws.cell(row=row, column=total_col).border = border_left
                ws.cell(row=row, column=total_col+1).value = '='
                ws.cell(row=row, column=total_col+1).font = font_bold
                ws.cell(row=row, column=total_col+1).number_format = '#,##0.00;[Red](#,##0.00)'
                ws.cell(row=row, column=total_col+2).value = '=' + get_column_letter(total_col) + str(row) + '+' + get_column_letter(total_col+1) + str(row)
                ws.cell(row=row, column=total_col+2).font = font_bold
                ws.cell(row=row, column=total_col+2).number_format = '#,##0.00'
            date = datetime.strptime(subscriber['date'], '%Y-%m-%d')
            if date in start_dates:
                col = (start_dates.index(date) + 1) * 2
                ws.cell(row=row, column=col).value = subscriber['value']
                ws.cell(row=row, column=col).number_format = '#,##0.00'
                ws.cell(row=row, column=total_col).value += '+' + get_column_letter(col) + str(row)
                self.env.cr.execute(sql_credits, (tuple(product_ids), subscriber['id'], date, end_dates[start_dates.index(date)],))
                credits = self.env.cr.dictfetchone()
                if credits != None:
                    if credits['value'] != None:
                        ws.cell(row=row, column=col+1).value = 0 - credits['value']
                    ws.cell(row=row, column=col+1).number_format = '#,##0.00;[Red](#,##0.00)'
                    ws.cell(row=row, column=total_col+1).value += '+' + get_column_letter(col+1) + str(row)
        final_row = row
        row += 1
        ws.cell(row=row, column=1).value = _('Total:')
        ws.cell(row=row, column=1).border = border_top_right
        ws.cell(row=row, column=1).font = font_bold
        for col in range(2, (total_col + 3)):
            letter = get_column_letter(col)
            ws.cell(row=row, column=col).value = '=SUM(' + letter + str(first_row) + ':' + letter + str(final_row) + ')'
            ws.cell(row=row, column=col).font = font_bold
            ws.cell(row=row, column=col).border = border_top_left if col == total_col else border_top
            if (col % 2) == 0:
                ws.cell(row=row, column=col).number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=col).number_format = '#,##0.00;[Red](#,##0.00)'

        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)
        file_name = _('Monthly subscriptions statement - ') + self.subscription_service_id.name + '.xlsx'

        self.write({
            'state': 'get',
            'name': file_name.encode('utf-8'),
            'data': out
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sbg.monthly.subscription.statement.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
