# -*- coding: utf-8 -*-
from openerp import models, fields, api, _
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
import base64
import datetime
import calendar
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment

class sbg_fiscal_general_ledger_wizard(models.TransientModel):
    _name = 'sbg.fiscal.general.ledger.wizard'
    _description = 'Fiscal general ledger wizard'

    # @api.multi
    # def _current_month(self):
    #
    name = fields.Char(string='Report name', default='Libro de diario mayor general')
    month = fields.Selection([
        (1, 'Enero'),
        (2, 'Febrero'),
        (3, 'Marzo'),
        (4, 'Abril'),
        (5, 'Mayo'),
        (6, 'Junio'),
        (7, 'Julio'),
        (8, 'Agosto'),
        (9, 'Septiembre'),
        (10, 'Octubre'),
        (11, 'Noviembre'),
        (12, 'Diciembre'),
    ], default=datetime.date.today().month)
    year = fields.Integer('Year', default=datetime.date.today().year)
    type = fields.Selection([('summary', 'Summary'), ('detailed', 'Detailed')], default='summary')
    first_page_number = fields.Integer('First page number', default=1)
    file_name = fields.Char('File Name', readonly=True)
    data = fields.Binary('File', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get'),
                              ('error', 'error')],
                             default='choose')

    @api.multi
    def back(self):
        self.state = 'choose'
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
        }

    @api.multi
    def generate(self):
        month = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        date_from = datetime.date(self.year, self.month, 1)
        date_to = datetime.date(self.year, self.month, calendar.monthrange(self.year, self.month)[1])
        is_summary = (self.type == 'summary')
        is_detail = (self.type == 'detailed')

        sql = """SELECT l.account_id, a.code, a.name,
                ROUND(SUM(CASE WHEN l.date < %(date_from)s THEN l.debit - l.credit ELSE 0 END)::NUMERIC, 2) as previous_balance,
                ROUND(SUM(CASE WHEN l.date BETWEEN %(date_from)s AND %(date_to)s THEN l.debit ELSE 0 END)::NUMERIC, 2) as debit,
                ROUND(SUM(CASE WHEN l.date BETWEEN %(date_from)s AND %(date_to)s THEN l.credit ELSE 0 END)::NUMERIC, 2) as credit
                FROM account_move_line l
                JOIN account_account a
                ON a.id = l.account_id
                JOIN account_move m
                ON m.id = l.move_id
                WHERE l.date <= %(date_to)s
                AND EXTRACT('year' FROM l.date) = %(year)s
                GROUP BY l.account_id, a.code, a.name
                ORDER BY a.code, a.name, l.account_id
        """

        if is_detail:
            sql_detail = """SELECT m.date, l.name as concept, m.name as document, l.debit, l.credit
                            FROM account_move_line l
                            JOIN account_account a
                            ON a.id = l.account_id
                            JOIN account_move m
                            ON m.id = l.move_id
                            WHERE l.date BETWEEN %(date_from)s AND %(date_to)s
                            AND l.account_id = %(account_id)s
                            ORDER BY a.code, a.name, l.account_id
            """

        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()

        border_top = Border(top=Side(style='thin'))
        border_bottom = Border(bottom=Side(style='thin'))
        font_bold = Font(bold=True)
        font_h1 = Font(size=18, bold=True)
        font_h2 = Font(size=14, bold=True)

        wb = Workbook()
        ws = wb.active

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.6
        ws.page_margins.right = 0.6

        ws.header_footer.left_header.text = 'Sociedad Biblica de Guatemala'
        ws.header_footer.left_header.font_size = 14
        ws.header_footer.left_header.font_name = 'Arial,Bold'
        ws.header_footer.center_header.text = '&BDIARIO MAYOR GENERAL&B'
        ws.header_footer.center_header.font_size = 18
        ws.header_footer.center_header.font_name = 'Arial,Bold'
        ws.header_footer.right_header.text = 'Hoja No. &P+' + str(self.first_page_number - 1)

        ws.title = "Diario mayor general"

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10 if is_detail else 0
        ws.column_dimensions['C'].width = 45 if is_detail else 54
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 13

        row = 1
        ws.cell(row=row, column=1).value = 'NIT: 1234567-9'
        ws.cell(row=row, column=1).font = font_bold
        ws.cell(row=row, column=4).value = 'Mes de ' + month[self.month] + ' de ' + str(self.year)
        ws.cell(row=row, column=4).font = font_bold
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')

        row += 1
        ws.cell(row=row, column=5).value = 'TRANSACCIONES DEL MES'
        ws.cell(row=row, column=5).font = font_bold
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')

        row += 1
        ws.cell(row=row, column=1).value = 'Cuenta'
        ws.cell(row=row, column=1).font = font_bold
        ws.cell(row=row, column=1).border = border_bottom
        ws.cell(row=row, column=2).value = 'Fecha' if is_detail else ''
        ws.cell(row=row, column=2).font = font_bold
        ws.cell(row=row, column=2).border = border_bottom
        ws.cell(row=row, column=3).value = 'Nombre de cuenta y concepto'
        ws.cell(row=row, column=3).font = font_bold
        ws.cell(row=row, column=3).border = border_bottom
        ws.cell(row=row, column=4).value = 'Documento' if is_detail else 'Saldo anterior'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='left' if is_detail else 'right')
        ws.cell(row=row, column=4).font = font_bold
        ws.cell(row=row, column=4).border = border_bottom
        ws.cell(row=row, column=5).value = 'Debe'
        ws.cell(row=row, column=5).font = font_bold
        ws.cell(row=row, column=5).border = border_bottom
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6).value = 'Haber'
        ws.cell(row=row, column=6).font = font_bold
        ws.cell(row=row, column=6).border = border_bottom
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=7).value = 'Saldo'
        ws.cell(row=row, column=7).font = font_bold
        ws.cell(row=row, column=7).border = border_bottom
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')

        ws.print_title_rows = '1:3'

        initial_row = row + 1

        self.env.cr.execute(sql, {'date_from': date_from, 'date_to': date_to, 'year': self.year})
        for line in self.env.cr.dictfetchall():
            if line["previous_balance"] != 0 or line["debit"] != 0 or line["debit"] != 0:
                row += 1
                ws.cell(row=row, column=1).value = line["code"]
                ws.cell(row=row, column=3).value = line["name"]
                ws.cell(row=row, column=3).style.alignment.wrap_text = True
                if is_summary:
                    ws.cell(row=row, column=4).value = line["previous_balance"]
                    ws.cell(row=row, column=4).number_format = '#,##0.00'
                    ws.cell(row=row, column=5).value = line["debit"]
                    ws.cell(row=row, column=6).value = line["credit"]
                    ws.cell(row=row, column=7).value = line["previous_balance"] + line["debit"] - line["credit"]
                    ws.cell(row=row, column=5).number_format = '#,##0.00'
                    ws.cell(row=row, column=6).number_format = '#,##0.00'
                    ws.cell(row=row, column=7).number_format = '#,##0.00'
                elif is_detail:
                    ws.cell(row=row, column=7).value = line["previous_balance"]
                    ws.cell(row=row, column=7).number_format = '#,##0.00'
                    self.env.cr.execute(sql_detail, {'date_from': date_from, 'date_to': date_to, 'account_id': line['account_id']})
                    row += 1
                    first_row = row
                    balance = line["previous_balance"]
                    for detail in self.env.cr.dictfetchall():
                        balance += detail["debit"] - detail["credit"]
                        ws.cell(row=row, column=2).value = datetime.datetime.strptime(detail["date"], '%Y-%m-%d')
                        ws.cell(row=row, column=3).value = detail["concept"]
                        ws.cell(row=row, column=4).value = detail["document"]
                        ws.cell(row=row, column=5).value = detail["debit"]
                        ws.cell(row=row, column=6).value = detail["credit"]
                        ws.cell(row=row, column=7).value = balance
                        ws.cell(row=row, column=2).number_format = 'Dd/Mm/yyyy'
                        ws.cell(row=row, column=5).number_format = '#,##0.00'
                        ws.cell(row=row, column=6).number_format = '#,##0.00'
                        ws.cell(row=row, column=7).number_format = '#,##0.00'
                        row += 1
                    if row == first_row:
                        row -= 1
                        ws.cell(row=row, column=5).value = line["debit"]
                        ws.cell(row=row, column=6).value = line["credit"]
                        ws.cell(row=row, column=7).value = line["previous_balance"] + line["debit"] - line["credit"]
                        ws.cell(row=row, column=5).number_format = '#,##0.00'
                        ws.cell(row=row, column=6).number_format = '#,##0.00'
                        ws.cell(row=row, column=7).number_format = '#,##0.00'
                    else:
                        ws.cell(row=row, column=4).value = 'Total:'
                        ws.cell(row=row, column=5).value = '=SUM(E{}:E{})'.format(first_row, row-1)
                        ws.cell(row=row, column=6).value = '=SUM(F{}:F{})'.format(first_row, row-1)
                        ws.cell(row=row, column=7).value = '={}+E{}-F{}'.format(line["previous_balance"], row, row)
                        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
                        ws.cell(row=row, column=5).number_format = '#,##0.00'
                        ws.cell(row=row, column=6).number_format = '#,##0.00'
                        ws.cell(row=row, column=7).number_format = '#,##0.00'
                    row += 1
        final_row = row
        row += 1
        if initial_row < final_row and is_summary:
            ws.cell(row=row, column=3).value = 'Total:'
            ws.cell(row=row, column=4).value = '=SUM(D{}:D{})'.format(initial_row, final_row)
            ws.cell(row=row, column=5).value = '=SUM(E{}:E{})'.format(initial_row, final_row)
            ws.cell(row=row, column=6).value = '=SUM(F{}:F{})'.format(initial_row, final_row)
            ws.cell(row=row, column=7).value = '=D{row}+E{row}-F{row}'.format(row=row)
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            for col in range(3, 8):
                if col > 3:
                    ws.cell(row=row, column=5).number_format = '#,##0.00'
                ws.cell(row=row, column=col).font = font_bold
                ws.cell(row=row, column=col).border = border_top

        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)
        file_name = '{} a {} de {} - {}.xlsx'.format(self.name, month[self.month], self.year, 'detallado' if is_detail else 'resumido')

        self.write({
            'state': 'get',
            'file_name': file_name,
            'data': out
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
