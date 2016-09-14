# -*- coding: utf-8 -*-
# from openerp.osv import fields, osv, orm


from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.worksheet import ColumnDimension
from openpyxl.styles import Border, Side, Font, Style
import base64
from openerp.tools.translate import _
from openerp import models, fields, api, exceptions, tools


class generate_stock_date_wizard(models.TransientModel):
    _name = "sbg.stock.date.wizard"
    date = fields.Date('Date', default=fields.Date.today)
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get')], default='choose')

    def generate_file(self, cr, uid, ids, context=None):
        def _get_column_letter(col_idx):
            """Convert a column number into a column letter (3 -> 'C')

            Right shift the column col_idx by 26 to find column letters in reverse
            order.  These numbers are 1-based, and can be converted to ASCII
            ordinals by adding 64.

            """
            # these indicies corrospond to A -> ZZZ and include all allowed
            # columns
            if not 1 <= col_idx <= 18278:
                raise ValueError("Invalid column index {0}".format(col_idx))
            letters = []
            while col_idx > 0:
                col_idx, remainder = divmod(col_idx, 26)
                # check for exact division and borrow if needed
                if remainder == 0:
                    remainder = 26
                    col_idx -= 1
                letters.append(chr(remainder + 64))
            return ''.join(reversed(letters))

        _COL_STRING_CACHE = {}
        _STRING_COL_CACHE = {}
        for col_cache in range(1, 18279):
            col_value = _get_column_letter(col_cache)
            _STRING_COL_CACHE[col_cache] = col_value
            _COL_STRING_CACHE[col_cache] = col_value

        def get_column_letter(idx, ):
            """Convert a column index into a column letter
            (3 -> 'C')
            """
            try:
                return _STRING_COL_CACHE[idx]
            except KeyError:
                raise ValueError("Invalid column index {0}".format(idx))


        if context is None:
            context = {}

        product_obj = self.pool.get('product.product')
        quant_obj = self.pool.get('stock.quant')
        location_obj = self.pool.get('stock.location')

        this = self.browse(cr, uid, ids)[0]
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()

        wb = Workbook()

        ws = wb.active

        title = _("Total stock at {}/{}/{}")
        ws.title = _("Total stock")
        ws['A1'].value = title.format(this.date[8:10], this.date[5:7], this.date[0:4])

        ws['A2'].value = _("id")
        ws['B2'].value = _("Ref")
        ws['C2'].value = _("EAN13")
        ws['D2'].value = _("Category")
        ws['E2'].value = _("Name")
        ws['F2'].value = _("Cost")
        ws['G2'].value = _("Price")
        ws['H2'].value = _("Stock")
        ws['I2'].value = _("Value")
        ws.merge_cells('A1:I1')
        #        ws.freeze_panes = 'A2'  no funciona


        border_bottom = Border(bottom=Side(style='thin'))
        border_right = Border(right=Side(style='thin'))
        border_corner = Border(right=Side(style='thin'), bottom=Side(style='thin'))
        font_bold = Font(bold=True)

        sql = """
            SELECT m.product_id, p.ean13, p.default_code, p.name_template, t.list_price, COALESCE(c.name, '') AS category,
            SUM(CASE WHEN ld.usage = 'internal' THEN m.product_qty ELSE 0 END - CASE WHEN ls.usage = 'internal' THEN m.product_qty ELSE 0 END) AS qty,
            mp.cost
                FROM stock_move m
                JOIN product_product p
                ON p.id = m.product_id
                JOIN stock_location ls
                ON ls.id = m.location_id
                JOIN stock_location ld
                ON ld.id = m.location_dest_id
                JOIN product_template t
                ON t.id = p.product_tmpl_id
                LEFT JOIN product_category c
                ON c.id = t.categ_id
                LEFT JOIN (
                    SELECT product_id, MAX(date) AS ultima_fecha
                    FROM stock_move
                    WHERE date <= %(fecha)s
                    GROUP BY product_id
                ) mu
                ON mu.product_id = m.product_id
                LEFT JOIN (
                    SELECT product_id, date, max(price_unit) AS cost FROM stock_move
                    WHERE date <= %(fecha)s
                    GROUP BY product_id, date
                ) mp
                ON mp.product_id = m.product_id
                AND mp.date = mu.ultima_fecha
                WHERE m.date <= %(fecha)s
                AND m.state = 'done'
                GROUP BY m.product_id, p.ean13, p.default_code, p.name_template, t.list_price, c.name, mp.cost
                HAVING SUM(CASE WHEN ld.usage = 'internal' THEN m.product_qty ELSE 0 END - CASE WHEN ls.usage = 'internal' THEN m.product_qty ELSE 0 END) <> 0
                ORDER BY p.ean13, m.product_id
        """
        params = {'fecha': this.date,}
        cr.execute(sql, {'fecha': this.date,})
        row = 3
        for product_line in cr.dictfetchall():
            # product_id = product_obj.browse(cr, uid, product_line['product_id'])
            cost = product_line['cost'] if product_line['cost'] else 0
            qty = product_line['qty'] if product_line['qty'] else 0

            ws.cell(row=row, column=1).value = product_line['product_id']
            ws.cell(row=row, column=2).value = product_line['default_code']
            ws.cell(row=row, column=3).value = product_line['ean13']
            ws.cell(row=row, column=4).value = product_line['category']
            ws.cell(row=row, column=5).value = product_line['name_template']
            ws.cell(row=row, column=6).value = cost
            ws.cell(row=row, column=7).value = product_line['list_price']
            ws.cell(row=row, column=8).value = qty
            ws.cell(row=row, column=9).value = '=F' + str(row) + '*H' + str(row)

            ws.cell(row=row, column=1).border = border_right
            ws.cell(row=row, column=2).border = border_right
            ws.cell(row=row, column=5).border = border_right
            ws.cell(row=row, column=6).border = border_right
            ws.cell(row=row, column=7).border = border_right
            ws.cell(row=row, column=8).border = border_right
            ws.cell(row=row, column=9).border = border_right
            row += 1

        for r in ws.iter_rows('A1:G1'):
            for c in r:
                c.font = font_bold
        for r in ws.iter_rows('A2:G2'):
            for c in r:
                c.border = border_bottom
                c.font = font_bold

        ws['A1'].border = border_right
        ws['A2'].border = border_corner
        ws['B1'].border = border_right
        ws['B2'].border = border_corner
        ws['C1'].border = border_right
        ws['C2'].border = border_corner
        ws['D1'].border = border_right
        ws['D2'].border = border_corner
        ws['E1'].border = border_right
        ws['E2'].border = border_corner
        ws['F1'].border = border_right
        ws['F2'].border = border_corner
        ws['G1'].border = border_right
        ws['G2'].border = border_corner
        ws['H1'].border = border_right
        ws['H2'].border = border_corner
        ws['I1'].border = border_right
        ws['I2'].border = border_corner

        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['E'].width = 70

        #
        # Location = Iternal
        #

        location_ids = location_obj.search(cr, uid, [('usage', '=', 'internal')], context=context)
        for location in location_ids:

            location_id = location_obj.browse(cr, uid, location)
            ws = wb.create_sheet()
            ws['A1'].value = location_id.display_name + ': {}/{}/{}'.format(this.date[8:10], this.date[5:7], this.date[0:4])
            ws.title = location_id.display_name.replace('/', '_')
            ws['A2'].value = _("id")
            ws['B2'].value = _("Ref")
            ws['C2'].value = _("EAN13")
            ws['D2'].value = _("Category")
            ws['E2'].value = _("Name")
            ws['F2'].value = _("Cost")
            ws['G2'].value = _("Price")
            ws['H2'].value = _("Stock")
            ws['I2'].value = _("Value")
            ws.merge_cells('A1:I1')

            sql = """
                SELECT m.product_id, p.ean13, p.default_code, p.name_template, t.list_price, COALESCE(c.name, '') AS category,
                SUM(CASE WHEN m.location_dest_id = %(bodega)s THEN m.product_qty ELSE 0 - m.product_qty END) AS qty,
                mp.cost
                    FROM stock_move m
                    JOIN product_product p
                    ON p.id = m.product_id
                    JOIN product_template t
                    ON t.id = p.product_tmpl_id
                    LEFT JOIN product_category c
                    ON c.id = t.categ_id
                    LEFT JOIN (
                        SELECT product_id, MAX(date) AS ultima_fecha
                        FROM stock_move
                        WHERE date <= %(fecha)s
                        AND (location_id = %(bodega)s OR location_dest_id = %(bodega)s)
                        GROUP BY product_id
                    ) mu
                    ON mu.product_id = m.product_id
                    LEFT JOIN (
                        SELECT product_id, date, max(price_unit) AS cost FROM stock_move
                        WHERE date <= %(fecha)s
                        AND (location_id = %(bodega)s OR location_dest_id = %(bodega)s)
                        GROUP BY product_id, date
                    ) mp
                    ON mp.product_id = m.product_id
                    AND mp.date = mu.ultima_fecha
                    WHERE (m.location_id = %(bodega)s OR m.location_dest_id = %(bodega)s)
                    AND m.date <= %(fecha)s
                    AND m.state = 'done'
                    GROUP BY m.product_id, p.ean13, p.default_code, p.name_template, t.list_price, c.name, mp.cost
                    HAVING SUM(CASE WHEN m.location_dest_id = %(bodega)s THEN m.product_qty ELSE 0 - m.product_qty END) <> 0
                    ORDER BY p.ean13, m.product_id
            """
            params = {'bodega': location, 'fecha': this.date,}
            cr.execute(sql, {'bodega': location, 'fecha': this.date,})

            row = 3
            for product_stock in cr.dictfetchall():
                # product_id = product_obj.browse(cr, uid, product_stock['product_id'])

                # se implemento un query a quant para que sea mas eficiente la extraccion de datos de la DB
                # product_quant_ids = quant_obj.search(cr, uid, [('location_id', '=', location),('product_id', '=', product_id.id)], context=context)
                # for quant in product_quant_ids:
                #    quant_id = quant_obj.browse(cr, uid,quant)
                #    product_stock+=quant_id.qty

                cost = product_stock['cost'] if product_stock['cost'] else 0
                qty = product_stock['qty'] if product_stock['qty'] else 0

                ws.cell(row=row, column=1).value = product_stock['product_id']
                ws.cell(row=row, column=2).value = product_stock['default_code']
                ws.cell(row=row, column=3).value = product_stock['ean13']
                ws.cell(row=row, column=4).value = product_stock['category']
                ws.cell(row=row, column=5).value = product_stock['name_template']
                ws.cell(row=row, column=6).value = product_stock['cost']
                ws.cell(row=row, column=7).value = product_stock['list_price']
                ws.cell(row=row, column=8).value = qty
                ws.cell(row=row, column=9).value = '=F' + str(row) + '*H' + str(row)

                ws.cell(row=row, column=1).border = border_right
                ws.cell(row=row, column=2).border = border_right
                ws.cell(row=row, column=5).border = border_right
                ws.cell(row=row, column=6).border = border_right
                ws.cell(row=row, column=7).border = border_right
                ws.cell(row=row, column=8).border = border_right
                ws.cell(row=row, column=9).border = border_right

                row += 1

            for r in ws.iter_rows('A1:I1'):
                for c in r:
                    c.font = font_bold
            for r in ws.iter_rows('A2:I2'):
                for c in r:
                    c.border = border_bottom
                    c.font = font_bold

            ws['A1'].border = border_right
            ws['A2'].border = border_corner
            ws['B1'].border = border_right
            ws['B2'].border = border_corner
            ws['C1'].border = border_right
            ws['C2'].border = border_corner
            ws['D1'].border = border_right
            ws['D2'].border = border_corner
            ws['E1'].border = border_right
            ws['E2'].border = border_corner
            ws['F1'].border = border_right
            ws['F2'].border = border_corner
            ws['G1'].border = border_right
            ws['G2'].border = border_corner
            ws['H1'].border = border_right
            ws['H2'].border = border_corner
            ws['I1'].border = border_right
            ws['I2'].border = border_corner
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['E'].width = 70

        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)
        name = _('Product stock by ')

        self.write(cr, uid, ids, {
            'state': 'get',
            'name': name + ' {}-{}-{}'.format(this.date[8:10], this.date[5:7], this.date[0:4]) + ".xlsx",
            'data': out
        }, context=context)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sbg.stock.date.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }

