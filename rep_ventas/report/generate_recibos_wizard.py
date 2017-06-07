# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, Font, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side


import base64
from openerp import models, fields

class GenerateStockWizard(models.Model):
    _name = "rep_ventas.generate_recibos"
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get')], default='choose')
    period_id=fields.Many2one('account.period',string=' ')
    desde = fields.Date('Desde')
    hasta = fields.Date('Hasta')


    def generate_file(self, cr, uid, ids, context=None):


        if context is None:
            context = {}

            


        this = self.browse(cr, uid, ids)[0]
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()
        thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))


        wb = Workbook()

        ws = wb.active

        ws.title = "recibos"
        ws['A1'].value = "SOCIEDAD BIBLICA DE GUATEMALA"
        ws['A2'].value = "REPORTE DE RECIBOS"
        ws.merge_cells('A1:N1') 
        ws.merge_cells('A2:N2') 



        ws['A4'].value = "DOCUMENTO"
        ws['B4'].value = "CODIGO"
        ws['C4'].value = "NIT"
        ws['D4'].value = "NOMBRE" 
        ws['E4'].value = "PROMOTOR"
        ws['F4'].value = "MONTO"
        ws['G4'].value = "FECHA"
        ws['H4'].value = "REFERENCIA"
        ws['I4'].value = "DEPOSITO"
        ws['J4'].value = "FECHA RECIBO"      
        ws['K4'].value = "PERIODO"
        ws['L4'].value = "APLICADO"
        ws['M4'].value = "ORIGEN"
        ws['N4'].value = "FACTURA"

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15


        for r in ws.iter_rows('A1:N4'):
            for c in r:
                c.font = c.font.copy(size=12, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')
        
        sql = """

SELECT docdate, 
         dia, "Mes", 
        Case When "Mes" in (1,2) Then 1
             When "Mes" in (3,4) Then 2
             When "Mes" in (5,6) Then 3
             When "Mes" in (7,8) Then 4
             When "Mes" in (9,10) Then 5
             When "Mes" in (11,12) Then 6
        End bimestre,

        Case When "Mes" in (1,2,3) Then 1
             When "Mes" in (4,5,6) Then 2
             When "Mes" in (7,8,9) Then 3
             When "Mes" in (10,11,12) Then 4
        End trimestre,
        Case When "Mes" in (1,2,3,4) Then 1
             When "Mes" in (5,6,7,8) Then 2
             When "Mes" in (9,10,11,12) Then 3
        End cuatrimestre,

        Case When "Mes" in (1,2,3,5,6) Then 1
             When "Mes" in (7,8,9,10,11,12) Then 2
        End semestre,
        anio,        
        "Cliente_Original" cliente, "Nombre_Original" nombre_cliente, "Promotor_Original" promotor_cliente,  
        categoria_cliente, "IDREGION" region, "IDDEPTO" departamento, "IDMUNI" municipio, "Promotor_Factura" promotor_factura, "CLASE" clase, 
        "CODIGO" producto, "DESCRIPCION" producto_descripcion, "UNIDADES" unidades, "PRECIO" precio, 
        "TOTAL_PRE" total_precio, "COSTO" costo, "TOTAL_COS" total_costo, origin
        
         FROM "SBG_facturacion_historico_detalle" where "CLASE" != 'KIT' anio in(%s,%s,%s)
order by anio,"Mes" ;
            """
#        period_mes = self.browse(cr, uid, ids)[0].period_id.code[0:2]
#        Period_anio = self.browse(cr, uid, ids)[0].period_id.code[3:7]

        cr.execute(sql,(this.desde,this.hasta,))

        row = 5
#        doc_count = 0
        for query_line in cr.dictfetchall():

            ws.cell(row=row, column=1).value  = query_line['number']
            ws.cell(row=row, column=2).value  = query_line['codigo']
            ws.cell(row=row, column=3).value  = query_line['nit']
            ws.cell(row=row, column=4).value  = query_line['nombre']
            ws.cell(row=row, column=5).value  = query_line['login']
            ws.cell(row=row, column=6).value  = query_line['amount']
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=7).value  = query_line['date']
            ws.cell(row=row, column=8).value  = query_line['reference']
            ws.cell(row=row, column=9).value  = query_line['name']
            ws.cell(row=row, column=10).value  = query_line['fecha_recibo']
            ws.cell(row=row, column=11).value  = query_line['code']
            ws.cell(row=row, column=12).value  = query_line['aplicado']
            ws.cell(row=row, column=12).number_format = '#,##0.00'
            ws.cell(row=row, column=13).value  = query_line['ref']
            ws.cell(row=row, column=14).value  = query_line['xfac']
            row += 1
        
      
        
        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)

        self.write(cr, uid, ids, {
            'state': 'get',
            'name': "rep_recibos_spreadsheet.xlsx",
            'data': out
        }, context=context)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'rep_ventas.generate_recibos',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
