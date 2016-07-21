# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, Font, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side


import base64
from openerp import models, fields

class GenerateStockWizard(models.Model):
    _name = "rep_ventas.generate_detalle"
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get')], default='choose')
    partner_id=fields.Many2one('res.partner',string=' ')

    def generate_file(self, cr, uid, ids, context=None):


        if context is None:
            context = {}

        this = self.browse(cr, uid, ids)[0]
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()
        thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))

        wb = Workbook()
        ws = wb.active
        ws.title = "detalle"
        ws['A1'].value = "SOCIEDAD BIBLICA DE GUATEMALA"
        ws['A2'].value = "REPORTE DE DETALLE DE FACTURACION POR CLIENTE"
        ws.merge_cells('A1:Z1') 
        ws.merge_cells('A2:Z2') 

        ws['A4'].value = "Documento"
        ws['B4'].value = "Diario"
        ws['C4'].value = "Fecha"
        ws['D4'].value = "Dia" 
        ws['E4'].value = "Mes"
        ws['F4'].value = "Año"
        ws['G4'].value = "Categoria"
        ws['H4'].value = "Region"
        ws['I4'].value = "Depto."
        ws['J4'].value = "Municipio"      
        ws['K4'].value = "Cliente"
        ws['L4'].value = "Nombre"
        ws['M4'].value = "Promotor"
        ws['N4'].value = "Tipo"
        ws['O4'].value = "Clase"
        ws['P4'].value = "Codigo"
        ws['Q4'].value = "Descripcion"
        ws['R4'].value = "Unidades"
        ws['S4'].value = "Precio"
        ws['T4'].value = "Total"
        ws['U4'].value = "Costo"
        ws['V4'].value = "Total Costo"
        ws['W4'].value = "Margen"
        ws['X4'].value = "Margen %"
        ws['Y4'].value = "Bodega"
        ws['Z4'].value = "Albaran"

        for r in ws.iter_rows('A1:Z4'):
            for c in r:
                c.font = c.font.copy(size=12, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')
        
        sql = """
            SELECT sopnumbe,
                   docid,
                   docdate,
                   dia,
                   "Mes" mes,
                   anio,
                   categoria_cliente,
                   "IDREGION" region,
                   "IDDEPTO" depto,
                   "IDMUNI" muni,
                   "Cliente_Factura" cliente,
                   "Nombre_Factura" nombre,
                   "Promotor_Factura" promotor,
                   tipo_producto,
                   "CLASE" clase,
                   "CODIGO" codigo,
                   "DESCRIPCION" descripcion,
                   "UNIDADES" unidades,
                   "PRECIO" precio,
                   "TOTAL_PRE" total,
                   "COSTO" costo,
                   "TOTAL_COS" total_costo,
                   "MARGEN" margen,
                   "MARGENp" margenp,
                   stock_location_name bodega,
                   partner_id albaran
            FROM "SBG_facturacion_historico_detalle"
            WHERE partner_id = %s
              """
#        period_mes = self.browse(cr, uid, ids)[0].period_id.code[0:2]
#        Period_anio = self.browse(cr, uid, ids)[0].period_id.code[3:7]
        Partner_id = self.browse(cr, uid, ids)[0].partner_id.id

        cr.execute(sql,(Partner_id,))

        row = 5
        doc_count = 0
        for query_line in cr.dictfetchall():

            ws.cell(row=row, column=1).value  = query_line['sopnumbe']
            ws.cell(row=row, column=2).value  = query_line['docid']
            ws.cell(row=row, column=3).value  = query_line['docdate']
            ws.cell(row=row, column=4).value  = query_line['dia']
            ws.cell(row=row, column=5).value  = query_line['mes']
            ws.cell(row=row, column=6).value  = query_line['anio']
            ws.cell(row=row, column=7).value  = query_line['categoria_cliente']
            ws.cell(row=row, column=8).value  = query_line['region']
            ws.cell(row=row, column=9).value  = query_line['depto']
            ws.cell(row=row, column=10).value  = query_line['muni']
            ws.cell(row=row, column=11).value  = query_line['cliente']
            ws.cell(row=row, column=12).value  = query_line['nombre']
            ws.cell(row=row, column=13).value  = query_line['promotor']
            ws.cell(row=row, column=14).value  = query_line['tipo_producto']
            ws.cell(row=row, column=15).value  = query_line['clase']
            ws.cell(row=row, column=16).value  = query_line['codigo']
            ws.cell(row=row, column=17).value  = query_line['descripcion']            
            ws.cell(row=row, column=18).value  = round(query_line['unidades'],0)
            ws.cell(row=row, column=19).value  = query_line['precio']
            ws.cell(row=row, column=20).value  = query_line['total']
            ws.cell(row=row, column=21).value  = query_line['costo']
            ws.cell(row=row, column=22).value  = query_line['total_costo']
            ws.cell(row=row, column=23).value  = query_line['margen']
            ws.cell(row=row, column=24).value  = query_line['margenp']
            ws.cell(row=row, column=25).value  = query_line['bodega']
            ws.cell(row=row, column=26).value  = query_line['albaran']
 
            row += 1
        
        
        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)

        self.write(cr, uid, ids, {
            'state': 'get',
            'name': "rep_detalle_spreadsheet.xlsx",
            'data': out
        }, context=context)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'rep_ventas.generate_detalle',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
