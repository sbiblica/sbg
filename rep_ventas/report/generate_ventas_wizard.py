# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, Font, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side


import base64
from openerp import models, fields

class GenerateStockWizard(models.Model):
    _name = "rep_ventas.generate_ventas"
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get')], default='choose')
    period_id=fields.Many2one('account.period',string=' ')

    def generate_file(self, cr, uid, ids, context=None):


        if context is None:
            context = {}

            


        this = self.browse(cr, uid, ids)[0]
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()
        thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))


        wb = Workbook()

        ws = wb.active

        ws.title = "rentabilidad"
        ws['A1'].value = "SOCIEDAD BIBLICA DE GUATEMALA"
        ws['A2'].value = "RESUMEN DE RENTABILIDAD"
        ws.merge_cells('A1:H1') 
        ws.merge_cells('A2:H2') 

        ws['A4'].value = "Año"
        ws['B4'].value = "Mes"
        ws['C4'].value = "Clase"
        ws['D4'].value = "Unidades" 
        ws['E4'].value = "Precio"
        ws['F4'].value = "Costo"
        ws['G4'].value = "Margen"
        ws['H4'].value = "%"
        
        for r in ws.iter_rows('A1:H4'):
            for c in r:
                c.font = c.font.copy(size=12, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')


        sql = """
            SELECT anio, "Mes" mes, "CLASE" clase, "CLAS_DESCRIPCION", sum(unidades) unidades, sum(precio) precio,                sum(total_precio)total_precio,
                   sum(costo) costo,
                   sum(Case When "CLASE" IN ('DON','ASA','BXM','SEV','INT','LEC','DON-PRO','INT-PRO','GAR') Then total_precio Else total_costo End) total_costo,
                   sum(Case When "CLASE" IN ('DON','ASA','BXM','SEV','INT','LEC','DON-PRO','INT-PRO','GAR') Then 0 Else margen End) margen,
                   sum(Case When "CLASE" = 'DESCUENTOS' Then 0
                           When "CLASE" IN ('DON','ASA','BXM','SEV','INT','LEC','DON-PRO','INT-PRO','GAR') Then 0
                           Else (margen / Case When total_precio = 0 Then 1 Else total_precio end ) * 100  End) margenp
              FROM "SBG_Resumen_Rentabilidad" where "CLASE" != 'KIT' and anio =  %s and "Mes"  = %s
            Group by anio, "Mes", "CLASE", "CLAS_DESCRIPCION"
            Order by "CLASE"
            """
        period_mes = self.browse(cr, uid, ids)[0].period_id.code[0:2]
        Period_anio = self.browse(cr, uid, ids)[0].period_id.code[3:7]

        cr.execute(sql,(Period_anio,period_mes,))

        #Totales
        Tunidades = 0
        Tprecio = 0
        Tcosto = 0
        Tmargen = 0
        Tporcentaje = 0

        row = 5
        doc_count = 0
        for query_line in cr.dictfetchall():

            margen_p = 0
            if query_line['margen'] >0:
                margen_p = round((query_line['margen']/query_line['total_precio'])*100,2)
                Tporcentaje += margen_p
                doc_count += 1

            ws.cell(row=row, column=1).value  = query_line['mes']
            ws.cell(row=row, column=2).value  = query_line['anio']
            ws.cell(row=row, column=3).value  = query_line['clase']
            ws.cell(row=row, column=4).value  = round(query_line['unidades'],0)
            ws.cell(row=row, column=5).value  = round(query_line['total_precio'],2)
            ws.cell(row=row, column=6).value  = round(query_line['total_costo'],2)
            ws.cell(row=row, column=7).value  = round(query_line['margen'],2)
            ws.cell(row=row, column=8).value  = margen_p
            
            Tunidades += query_line['unidades']
            Tprecio   += query_line['total_precio']
            Tcosto    += query_line['total_costo']
            Tmargen   += query_line['margen']

            row += 1
       
        ws.cell(row=row, column=3).value  = "Total:"
        ws.cell(row=row, column=4).value  = round(Tunidades,0)
        ws.cell(row=row, column=5).value  = round(Tprecio,2)
        ws.cell(row=row, column=6).value  = round(Tcosto,2)
        ws.cell(row=row, column=7).value  = round(Tmargen,2)
        ws.cell(row=row, column=8).value  = round((margen_p/doc_count),2)


       #REPORTE DE VENTAS --> FACTURAS POR PERIODO
        ws = wb.create_sheet()
        ws.title = "ventas"
        ws['A1'].value = "SOCIEDAD BIBLICA DE GUATEMALA"
        ws['A2'].value = "REPORTE DE FACTURACION"
        ws.merge_cells('A1:O1') 
        ws.merge_cells('A2:O2') 

        ws['A4'].value = "Documento"
        ws['B4'].value = "Fecha"
        ws['C4'].value = "Dia"
        ws['D4'].value = "Mes" 
        ws['E4'].value = "Año"
        ws['F4'].value = "Cliente"
        ws['G4'].value = "Nombre"
        ws['H4'].value = "Estado"
        ws['I4'].value = "Promotor"
        ws['J4'].value = "Club"      
        ws['L4'].value = "Sin IVA"
        ws['M4'].value = "IVA"
        ws['N4'].value = "Total"
        ws['O4'].value = "Saldo"

        for r in ws.iter_rows('A1:O4'):
            for c in r:
                c.font = c.font.copy(size=12, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')
        
        sql = """
                SELECT
                 SBG_ventas."factura" AS documento,
                 SBG_ventas."fecha" AS fecha,
                 SBG_ventas."dia" AS dia,
                 SBG_ventas."mes" AS mes,
                 SBG_ventas."anio" AS anio,
                 SBG_ventas."cliente" AS cliente,
                 SBG_ventas."nombre" AS nombre,
                 SBG_ventas."sin_iva" AS sin_iva,
                 SBG_ventas."iva" AS iva,
                 SBG_ventas."total" AS total,
                 SBG_ventas."saldo" AS saldo,
                 SBG_ventas."estado" AS estado,
                 SBG_ventas."promotor" promotor,
                 SBG_ventas."diario" AS diario,
                 tipo.club as club
            FROM "SBG_facturacion_historico" SBG_ventas
                 LEFT JOIN
                 (SELECT c.sopnumbe sopnumbe,min('Club') club
                    FROM "SBG_facturacion_historico_detalle" c
                    where c."CLASE" in ('BXM','SEV','LEC')
                 group by c.sopnumbe) tipo
                 ON SBG_ventas."factura" = tipo.sopnumbe
            WHERE  SBG_ventas."anio" = %s and  SBG_ventas."mes" = %s
            
            ORDER BY
                 SBG_ventas."fecha" ASC
            """
        period_mes = self.browse(cr, uid, ids)[0].period_id.code[0:2]
        Period_anio = self.browse(cr, uid, ids)[0].period_id.code[3:7]

        cr.execute(sql,(Period_anio,period_mes,))

        row = 5
        doc_count = 0
        for query_line in cr.dictfetchall():

            ws.cell(row=row, column=1).value  = query_line['documento']
            ws.cell(row=row, column=2).value  = query_line['fecha']
            ws.cell(row=row, column=3).value  = query_line['dia']
            ws.cell(row=row, column=4).value  = query_line['mes']
            ws.cell(row=row, column=5).value  = query_line['anio']
            ws.cell(row=row, column=6).value  = query_line['cliente']
            ws.cell(row=row, column=7).value  = query_line['nombre']
            ws.cell(row=row, column=8).value  = query_line['estado']
            ws.cell(row=row, column=9).value  = query_line['promotor']
            ws.cell(row=row, column=10).value  = query_line['diario']
            ws.cell(row=row, column=11).value  = query_line['club']
            ws.cell(row=row, column=12).value  = round(query_line['sin_iva'],0)
            ws.cell(row=row, column=13).value  = round(query_line['iva'],2)
            ws.cell(row=row, column=14).value  = round(query_line['total'],2)
            ws.cell(row=row, column=15).value  = round(query_line['saldo'],2)
            row += 1
        
        


       #DETALLE DE VENTAS
        ws = wb.create_sheet()
        ws.title = "detalle"
        ws['A1'].value = "SOCIEDAD BIBLICA DE GUATEMALA"
        ws['A2'].value = "REPORTE DE DETALLE DE FACTURACION"
        ws.merge_cells('A1:Z1') 
        ws.merge_cells('A2:Z2') 

        ws['A4'].value = "Documento"
        ws['B4'].value = "Diario"
        ws['C4'].value = "Fecha"
        ws['D4'].value = "Dia" 
        ws['E4'].value = "Mes"
        ws['F4'].value = "Año"
        ws['G4'].value = "Categoria"
        ws['H4'].value = "Region"
        ws['I4'].value = "Depto."
        ws['J4'].value = "Municipio"      
        ws['K4'].value = "Cliente"
        ws['L4'].value = "Nombre"
        ws['M4'].value = "Promotor"
        ws['N4'].value = "Tipo"
        ws['O4'].value = "Clase"
        ws['P4'].value = "Codigo"
        ws['Q4'].value = "Descripcion"
        ws['R4'].value = "Unidades"
        ws['S4'].value = "Precio"
        ws['T4'].value = "Total"
        ws['U4'].value = "Costo"
        ws['V4'].value = "Total Costo"
        ws['W4'].value = "Margen"
        ws['X4'].value = "Margen %"
        ws['Y4'].value = "Bodega"
        ws['Z4'].value = "Albaran"

        for r in ws.iter_rows('A1:Z4'):
            for c in r:
                c.font = c.font.copy(size=12, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')
        
        sql = """
            SELECT sopnumbe,
                   docid,
                   docdate,
                   dia,
                   "Mes" mes,
                   anio,
                   categoria_cliente,
                   "IDREGION" region,
                   "IDDEPTO" depto,
                   "IDMUNI" muni,
                   "Cliente_Factura" cliente,
                   "Nombre_Factura" nombre,
                   "Promotor_Factura" promotor,
                   tipo_producto,
                   "CLASE" clase,
                   "CODIGO" codigo,
                   "DESCRIPCION" descripcion,
                   "UNIDADES" unidades,
                   "PRECIO" precio,
                   "TOTAL_PRE" total,
                   "COSTO" costo,
                   "TOTAL_COS" total_costo,
                   "MARGEN" margen,
                   "MARGENp" margenp,
                   stock_location_name bodega,
                   origin albaran
            FROM "SBG_facturacion_historico_detalle"
            WHERE anio = %s
              AND "Mes" = %s
              """
        period_mes = self.browse(cr, uid, ids)[0].period_id.code[0:2]
        Period_anio = self.browse(cr, uid, ids)[0].period_id.code[3:7]

        cr.execute(sql,(Period_anio,period_mes,))

        row = 5
        doc_count = 0
        for query_line in cr.dictfetchall():

            ws.cell(row=row, column=1).value  = query_line['sopnumbe']
            ws.cell(row=row, column=2).value  = query_line['docid']
            ws.cell(row=row, column=3).value  = query_line['docdate']
            ws.cell(row=row, column=4).value  = query_line['dia']
            ws.cell(row=row, column=5).value  = query_line['mes']
            ws.cell(row=row, column=6).value  = query_line['anio']
            ws.cell(row=row, column=7).value  = query_line['categoria_cliente']
            ws.cell(row=row, column=8).value  = query_line['region']
            ws.cell(row=row, column=9).value  = query_line['depto']
            ws.cell(row=row, column=10).value  = query_line['muni']
            ws.cell(row=row, column=11).value  = query_line['cliente']
            ws.cell(row=row, column=12).value  = query_line['nombre']
            ws.cell(row=row, column=13).value  = query_line['promotor']
            ws.cell(row=row, column=14).value  = query_line['tipo_producto']
            ws.cell(row=row, column=15).value  = query_line['clase']
            ws.cell(row=row, column=16).value  = query_line['codigo']
            ws.cell(row=row, column=17).value  = query_line['descripcion']            
            ws.cell(row=row, column=18).value  = round(query_line['unidades'],0)
            ws.cell(row=row, column=19).value  = query_line['precio']
            ws.cell(row=row, column=20).value  = query_line['total']
            ws.cell(row=row, column=21).value  = query_line['costo']
            ws.cell(row=row, column=22).value  = query_line['total_costo']
            ws.cell(row=row, column=23).value  = query_line['margen']
            ws.cell(row=row, column=24).value  = query_line['margenp']
            ws.cell(row=row, column=25).value  = query_line['bodega']
            ws.cell(row=row, column=26).value  = query_line['albaran']
 
            row += 1
        
        
        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)

        self.write(cr, uid, ids, {
            'state': 'get',
            'name': "rep_ventas_spreadsheet.xlsx",
            'data': out
        }, context=context)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'rep_ventas.generate_ventas',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
