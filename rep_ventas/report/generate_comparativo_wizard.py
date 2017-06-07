# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, Font, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side


import base64
from openerp import models, fields

class GenerateStockWizard(models.Model):
    _name = "rep_ventas.generate_comparativo"
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get')], default='choose')
    period_id=fields.Many2one('account.period',string=' ')
    ano1 = fields.Integer('Año 1')
    ano2 = fields.Integer('Año 2')
    ano3 = fields.Integer('Año 3')

    def generate_file(self, cr, uid, ids, context=None):


        if context is None:
            context = {}

            


        this = self.browse(cr, uid, ids)[0]
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()
        thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))


        wb = Workbook()

        ws = wb.active

        ws.title = "comparativo"
        ws['A1'].value = "SOCIEDAD BIBLICA DE GUATEMALA"
        ws['A2'].value = "REPORTE DE COMPARATIVO 3 AÑOS"
        ws.merge_cells('A1:X1') 
        ws.merge_cells('A2:X2') 

        ws['A4'].value = "DIA"
        ws['B4'].value = "MES"
        ws['C4'].value = "BIMESTRE"
        ws['D4'].value = "TRIMESTRE" 
        ws['E4'].value = "CUATRIMESTRE"
        ws['F4'].value = "SEMESTRE"
        ws['G4'].value = "AÑO"
        ws['H4'].value = "CLIENTE"
        ws['I4'].value = "NOMBRE DEL CLIENTE"
        ws['J4'].value = "PROMOTOR"      
        ws['K4'].value = "CATEGORIA"
        ws['L4'].value = "REGION"
        ws['M4'].value = "DEPARTAMENTO"
        ws['N4'].value = "MUNICIPIO"
        ws['O4'].value = "PROMOTOR FACTURA"
        ws['P4'].value = "CLASE"
        ws['Q4'].value = "PRODUCTO"
        ws['R4'].value = "DESCRIPCION"
        ws['S4'].value = "UNIDADES"
        ws['T4'].value = "PRECIO"
        ws['U4'].value = "TOTAL PRECIO"
        ws['V4'].value = "COSTO"
        ws['W4'].value = "TOTAL COSTO"
        ws['X4'].value = "ORIGEN"

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 40
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 20
        ws.column_dimensions['P'].width = 20
        ws.column_dimensions['Q'].width = 20
        ws.column_dimensions['R'].width = 40
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15
        ws.column_dimensions['X'].width = 15


        for r in ws.iter_rows('A1:X4'):
            for c in r:
                c.font = c.font.copy(size=12, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')
        
        sql = """

SELECT docdate, 
         dia, "Mes" mes, 
        Case When "Mes" in (1,2) Then 1
             When "Mes" in (3,4) Then 2
             When "Mes" in (5,6) Then 3
             When "Mes" in (7,8) Then 4
             When "Mes" in (9,10) Then 5
             When "Mes" in (11,12) Then 6
        End bimestre,

        Case When "Mes" in (1,2,3) Then 1
             When "Mes" in (4,5,6) Then 2
             When "Mes" in (7,8,9) Then 3
             When "Mes" in (10,11,12) Then 4
        End trimestre,
        Case When "Mes" in (1,2,3,4) Then 1
             When "Mes" in (5,6,7,8) Then 2
             When "Mes" in (9,10,11,12) Then 3
        End cuatrimestre,

        Case When "Mes" in (1,2,3,5,6) Then 1
             When "Mes" in (7,8,9,10,11,12) Then 2
        End semestre,
        anio,        
        "Cliente_Original" cliente, "Nombre_Original" nombre_cliente, "Promotor_Original" promotor_cliente,  
        categoria_cliente, "IDREGION" region, "IDDEPTO" departamento, "IDMUNI" municipio, "Promotor_Factura" promotor_factura, "CLASE" clase, 
            "CODIGO" producto, "DESCRIPCION" producto_descripcion, "UNIDADES" unidades, "PRECIO" precio, 
        "TOTAL_PRE" total_precio, "COSTO" costo, "TOTAL_COS" total_costo, origin
         FROM "SBG_facturacion_historico_detalle" where "CLASE" != 'KIT' and anio in(%s,%s,%s);
         """
#        period_mes = self.browse(cr, uid, ids)[0].period_id.code[0:2]
#        Period_anio = self.browse(cr, uid, ids)[0].period_id.code[3:7]

        cr.execute(sql,(this.ano1,this.ano2,this.ano3,))

        row = 5
#        doc_count = 0
        for query_line in cr.dictfetchall():

            ws.cell(row=row, column=1).value  = query_line['dia']
            ws.cell(row=row, column=2).value  = query_line['mes']
            ws.cell(row=row, column=3).value  = query_line['bimestre']
            ws.cell(row=row, column=4).value  = query_line['trimestre']
            ws.cell(row=row, column=5).value  = query_line['cuatrimestre']
            ws.cell(row=row, column=6).value  = query_line['semestre']
            ws.cell(row=row, column=7).value  = query_line['anio']
            ws.cell(row=row, column=8).value  = query_line['cliente']
            ws.cell(row=row, column=9).value  = query_line['nombre_cliente']
            ws.cell(row=row, column=10).value  = query_line['promotor_cliente']
            ws.cell(row=row, column=11).value  = query_line['categoria_cliente']
            ws.cell(row=row, column=12).value  = query_line['region']
            ws.cell(row=row, column=13).value  = query_line['departamento']
            ws.cell(row=row, column=14).value  = query_line['municipio']
            ws.cell(row=row, column=15).value  = query_line['promotor_factura']
            ws.cell(row=row, column=16).value  = query_line['clase']
            ws.cell(row=row, column=17).value  = query_line['producto']
            ws.cell(row=row, column=18).value  = query_line['producto_descripcion']
            ws.cell(row=row, column=19).value  = query_line['unidades']
            ws.cell(row=row, column=19).number_format = '#,##0'
            ws.cell(row=row, column=20).value  = query_line['precio']
            ws.cell(row=row, column=20).number_format = '#,##0.00'
            ws.cell(row=row, column=21).value  = query_line['total_precio']
            ws.cell(row=row, column=21).number_format = '#,##0.00'
            ws.cell(row=row, column=22).value  = query_line['costo']
            ws.cell(row=row, column=22).number_format = '#,##0.00'
            ws.cell(row=row, column=23).value  = query_line['total_costo']
            ws.cell(row=row, column=23).number_format = '#,##0.00'
            ws.cell(row=row, column=24).value  = query_line['origin']
            ws.cell(row=row, column=24).number_format = '#,##0.00'

            row += 1
        
      
        
        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)

        self.write(cr, uid, ids, {
            'state': 'get',
            'name': "rep_comparativo_spreadsheet.xlsx",
            'data': out
        }, context=context)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'rep_ventas.generate_comparativo',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
