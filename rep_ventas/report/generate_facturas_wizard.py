# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, Font, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side


import base64
from openerp import models, fields

class GenerateStockWizard(models.Model):
    _name = "rep_ventas.generate_facturas"
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get')], default='choose')
    period_id=fields.Many2one('account.period',string=' ')
    desde = fields.Date('Desde')
    hasta = fields.Date('Hasta')


    def generate_file(self, cr, uid, ids, context=None):


        if context is None:
            context = {}

            


        this = self.browse(cr, uid, ids)[0]
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()
        thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))


        wb = Workbook()

        ws = wb.active

        ws.title = "facturas"
        ws['A1'].value = "SOCIEDAD BIBLICA DE GUATEMALA"
        ws['A2'].value = "REPORTE DE FACTURACION"
        ws.merge_cells('A1:O1') 
        ws.merge_cells('A2:O2') 

        ws['A4'].value = "Documento"
        ws['B4'].value = "Fecha"
        ws['C4'].value = "Dia"
        ws['D4'].value = "Mes" 
        ws['E4'].value = "Año"
        ws['F4'].value = "Cliente"
        ws['G4'].value = "Nombre"
        ws['H4'].value = "Estado"
        ws['I4'].value = "Promotor"
        ws['J4'].value = "Club"      
        ws['L4'].value = "Sin IVA"
        ws['M4'].value = "IVA"
        ws['N4'].value = "Total"
        ws['O4'].value = "Saldo"

        for r in ws.iter_rows('A1:O4'):
            for c in r:
                c.font = c.font.copy(size=12, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')
        
        sql = """
                SELECT
                 SBG_ventas."factura" AS documento,
                 SBG_ventas."fecha" AS fecha,
                 SBG_ventas."dia" AS dia,
                 SBG_ventas."mes" AS mes,
                 SBG_ventas."anio" AS anio,
                 SBG_ventas."cliente" AS cliente,
                 SBG_ventas."nombre" AS nombre,
                 SBG_ventas."sin_iva" AS sin_iva,
                 SBG_ventas."iva" AS iva,
                 SBG_ventas."total" AS total,
                 SBG_ventas."saldo" AS saldo,
                 SBG_ventas."estado" AS estado,
                 SBG_ventas."promotor" promotor,
                 SBG_ventas."diario" AS diario,
                 tipo.club as club
            FROM "SBG_facturacion_historico" SBG_ventas
                 LEFT JOIN
                 (SELECT c.sopnumbe sopnumbe,min('Club') club
                    FROM "SBG_facturacion_historico_detalle" c
                    where c."CLASE" in ('BXM','SEV','LEC')
                 group by c.sopnumbe) tipo
                 ON SBG_ventas."factura" = tipo.sopnumbe
            WHERE  SBG_ventas."fecha" between %s and %s
            
            ORDER BY
                 SBG_ventas."fecha" ASC
            """
#        period_mes = self.browse(cr, uid, ids)[0].period_id.code[0:2]
#        Period_anio = self.browse(cr, uid, ids)[0].period_id.code[3:7]

        cr.execute(sql,(this.desde,this.hasta,))

        row = 5
#        doc_count = 0
        for query_line in cr.dictfetchall():

            ws.cell(row=row, column=1).value  = query_line['documento']
            ws.cell(row=row, column=2).value  = query_line['fecha']
            ws.cell(row=row, column=3).value  = query_line['dia']
            ws.cell(row=row, column=4).value  = query_line['mes']
            ws.cell(row=row, column=5).value  = query_line['anio']
            ws.cell(row=row, column=6).value  = query_line['cliente']
            ws.cell(row=row, column=7).value  = query_line['nombre']
            ws.cell(row=row, column=8).value  = query_line['estado']
            ws.cell(row=row, column=9).value  = query_line['promotor']
            ws.cell(row=row, column=10).value  = query_line['diario']
            ws.cell(row=row, column=11).value  = query_line['club']
            ws.cell(row=row, column=12).value  = round(query_line['sin_iva'],0)
            ws.cell(row=row, column=13).value  = round(query_line['iva'],2)
            ws.cell(row=row, column=14).value  = round(query_line['total'],2)
            ws.cell(row=row, column=15).value  = round(query_line['saldo'],2)
            row += 1
        
        
        
        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)

        self.write(cr, uid, ids, {
            'state': 'get',
            'name': "rep_ventas_spreadsheet.xlsx",
            'data': out
        }, context=context)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'rep_ventas.generate_facturas',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
