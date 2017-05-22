# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, Font, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side


import base64
from openerp import models, fields

class GenerateStockWizard(models.Model):
    _name = "sbg_centro_costos.generate_centro_costos"
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get')], default='choose')
    anio = fields.Char('')
    account_id=fields.Many2one('account.analytic.account',string=' ')


    def generate_file(self, cr, uid, ids, context=None):


        if context is None:
            context = {}

        this = self.browse(cr, uid, ids)[0]
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()
        thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))


        wb = Workbook()

        ws = wb.active

        ws.title = "rentabilidad"
        ws['A1'].value = "SOCIEDAD BIBLICA ARGENTINA"
        ws['A2'].value = "REPORTE DE CENTROS DE COSTO"
        ws.merge_cells('A1:E1') 
        ws.merge_cells('A2:E2') 


        ws['A4'].value = "AÑO"
        ws['B4'].value = "PERIODO"
        ws['C4'].value = "CUENTA"
        ws['D4'].value = "NOMBRE" 
        ws['E4'].value = "MONTO"
        
        for r in ws.iter_rows('A1:E4'):
            for c in r:
                c.font = c.font.copy(size=12, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')


        sql = """

Select aal.account_id,af.name anio,aal.code periodo,aaa.code centro,aaa.name centro_costo,aa.code cuenta,aa.name nombre_cuenta,sum((aal.amount*-1)) valor
    from account_analytic_line aal
         JOIN account_analytic_account aaa ON aaa.id = aal.account_id
         JOIN account_account aa ON aa.id = aal.general_account_id
         JOIN account_period ap ON ap.code = aal.code
         JOIN account_fiscalyear af ON ap.fiscalyear_id = af.id
         where af.name = %s and aal.account_id = %s
Group by aal.account_id,af.name ,aal.code ,aaa.code ,aaa.name ,aa.code ,aa.name 
            """
#        period_mes = self.browse(cr, uid, ids)[0].period_id.code
#        Period_anio = self.browse(cr, uid, ids)[0].period_id.code[3:7]

        cr.execute(sql,(this.anio,this.account_id.id))

        #Totales
        Tmonto = 0

        row = 5
#        doc_count = 0
        for query_line in cr.dictfetchall():

            ws.cell(row=row, column=1).value  = query_line['anio']
            ws.cell(row=row, column=2).value  = query_line['periodo']
            ws.cell(row=row, column=3).value  = query_line['cuenta']
            ws.cell(row=row, column=4).value  = query_line['nombre_cuenta']
            ws.cell(row=row, column=5).value  = round(query_line['valor'],2)

            row += 1
       

        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)

        self.write(cr, uid, ids, {
            'state': 'get',
            'name': "rep_centro_costos_spreadsheet.xlsx",
            'data': out
        }, context=context)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sbg_centro_costos.generate_centro_costos',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
